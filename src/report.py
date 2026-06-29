"""판정 결과를 엑셀 리포트로 출력한다 (openpyxl).

- audit_report.xlsx : 판정결과 시트 + 요약 시트
- source_data.xlsx  : 원본 거래 데이터 (발주/입고/청구 시트)
"""
import openpyxl
from openpyxl.styles import PatternFill, Font


class ReportGenerator:
    """MatchResult 모음을 받아 엑셀 감사 리포트를 만든다."""

    HOLD_FILL = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    HEADER_FONT = Font(bold=True)

    def __init__(self, results, invoices, vendor_by_code):
        self.results = results
        self.invoices = invoices
        self.vendor_by_code = vendor_by_code

    def generate(self, path="output/audit_report.xlsx") -> str:
        """판정 결과 엑셀을 만들어 저장한다."""
        import os
        os.makedirs(os.path.dirname(path), exist_ok=True)
        wb = openpyxl.Workbook()
        self._add_result_sheet(wb)
        self._add_summary_sheet(wb)
        wb.save(path)
        return path

    def _add_result_sheet(self, wb):
        ws = wb.active
        ws.title = "판정결과"
        ws.append(["청구번호", "거래처", "상태", "사유"])
        for cell in ws[1]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT

        for invoice in self.invoices:
            mr = self.results[invoice.invoice_number]
            vendor = self.vendor_by_code[invoice.vendor_code]
            ws.append([
                mr.invoice_number,
                vendor.name,
                mr.status,
                ", ".join(mr.reasons) if mr.reasons else "-",
            ])
            if mr.status == "보류":
                for cell in ws[ws.max_row]:
                    cell.fill = self.HOLD_FILL

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 45

    def _add_summary_sheet(self, wb):
        ws = wb.create_sheet("요약")
        total = len(self.invoices)
        normal = sum(1 for inv in self.invoices if self.results[inv.invoice_number].status == "정상")
        hold = total - normal

        ws.append(["구분", "건수"])
        ws.append(["총 검증", total])
        ws.append(["정상", normal])
        ws.append(["보류", hold])
        ws.append([])

        sub_header_row = ws.max_row + 1
        ws.append(["사유별 분포", "건수"])
        reason_count = {}
        for inv in self.invoices:
            for check in self.results[inv.invoice_number].checks:
                if not check.passed:
                    reason_count[check.check_type] = reason_count.get(check.check_type, 0) + 1
        for key, cnt in reason_count.items():
            ws.append([key, cnt])

        for cell in ws[1]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
        for cell in ws[sub_header_row]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 8

    def generate_source(self, pos, grs, invs, path="output/source_data.xlsx") -> str:
        """원본 거래 데이터(발주/입고/청구)를 3개 시트로 출력한다."""
        import os
        os.makedirs(os.path.dirname(path), exist_ok=True)
        wb = openpyxl.Workbook()

        ws_po = wb.active
        ws_po.title = "발주(PO)"
        ws_po.append(["발주번호", "거래처코드", "발주일", "품목", "수량", "단가", "금액"])
        for po in pos:
            for item in po.items:
                ws_po.append([po.po_number, po.vendor_code, po.order_date,
                               item.material_code, item.quantity, item.unit_price, item.amount])

        ws_gr = wb.create_sheet("입고(GR)")
        ws_gr.append(["입고번호", "발주번호", "입고일", "품목", "수량"])
        for gr in grs:
            for item in gr.items:
                ws_gr.append([gr.gr_number, gr.po_number, gr.receipt_date,
                               item.material_code, item.quantity])

        ws_inv = wb.create_sheet("청구(Invoice)")
        ws_inv.append(["청구번호", "거래처코드", "청구일", "발주번호들", "품목", "수량", "단가", "금액"])
        for inv in invs:
            po_text = ", ".join(inv.po_numbers)
            for item in inv.items:
                ws_inv.append([inv.invoice_number, inv.vendor_code, inv.invoice_date,
                                po_text, item.material_code, item.quantity, item.unit_price, item.amount])

        for ws in [ws_po, ws_gr, ws_inv]:
            for cell in ws[1]:
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT

        wb.save(path)
        return path

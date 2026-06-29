"""
purchase-fraud-detector 레포 자동 구성 스크립트

사용법:
1. `purchase-fraud-detector` 레포를 로컬에 clone
   git clone https://github.com/dabok0519/purchase-fraud-detector.git
   cd purchase-fraud-detector

2. 이 스크립트를 레포 루트에 복사 후 실행
   python setup_repo.py

3. 완료 후 스크립트 파일 삭제
   del setup_repo.py  (Windows)
"""

import os
import subprocess

def run(cmd):
    print(f"  $ {cmd}")
    subprocess.run(cmd, shell=True, check=True)

def write(path, content):
    os.makedirs(os.path.dirname(path) if os.path.dirname(path) else ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)

def commit(msg):
    run("git add -A")
    run(f'git commit -m "{msg}"')


# ─────────────────────────────────────────────
# Commit 1: chore - init project structure
# ─────────────────────────────────────────────
print("\n[1/12] chore: init project structure")

write(".gitignore", """\
# 환경변수 (API 키)
.env

# 산출물
output/
*.xlsx

# Python
__pycache__/
*.py[cod]
*.egg-info/
dist/
build/
.venv/
venv/

# IDE
.vscode/
.idea/
""")

write(".env.example", """\
# 국세청 사업자등록 상태조회 API 인증키
# 공공데이터포털(https://www.data.go.kr)에서 발급
NTS_SERVICE_KEY=your_decoding_key_here
""")

write("requirements.txt", """\
requests
python-dotenv
openpyxl
""")

write("tests/.gitkeep", "")
write("output/.gitignore", """\
# 생성된 리포트 파일은 버전 관리하지 않음
*
!.gitignore
""")

commit("chore: init project structure")


# ─────────────────────────────────────────────
# Commit 2: feat - define core data models
# ─────────────────────────────────────────────
print("\n[2/12] feat: define core data models")

write("src/__init__.py", "")
write("src/models.py", """\
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal


@dataclass
class Material:
    \"\"\"자재 마스터. 품목 기본 정보를 담는 그릇.\"\"\"
    material_code: str
    name: str


@dataclass
class Vendor:
    \"\"\"거래처 마스터.\"\"\"
    vendor_code: str
    business_number: str
    name: str
    status: str = "미확인"

    VALID_STATUSES = ["정상", "폐업", "휴업", "검증불가", "미확인"]

    def __post_init__(self):
        if self.status not in self.VALID_STATUSES:
            raise ValueError(f"잘못된 status 값: {self.status}")


@dataclass
class POItem:
    \"\"\"발주서 내의 개별 라인 아이템.\"\"\"
    material_code: str
    quantity: Decimal
    unit_price: Decimal

    @property
    def amount(self) -> Decimal:
        return self.quantity * self.unit_price


@dataclass
class PurchaseOrder:
    \"\"\"발주 헤더 + 라인 모델.\"\"\"
    po_number: str
    vendor_code: str
    order_date: date
    items: list[POItem] = field(default_factory=list)


@dataclass
class GRItem:
    \"\"\"입고 라인. 실제 입고 수량을 기록.\"\"\"
    material_code: str
    quantity: Decimal


@dataclass
class GoodsReceipt:
    \"\"\"입고 헤더 + 라인 모델.\"\"\"
    gr_number: str
    po_number: str
    receipt_date: date
    items: list[GRItem] = field(default_factory=list)


@dataclass
class InvoiceItem:
    \"\"\"청구 라인. 실제 청구 수량과 단가를 기록.\"\"\"
    material_code: str
    quantity: Decimal
    unit_price: Decimal

    @property
    def amount(self) -> Decimal:
        return self.quantity * self.unit_price


@dataclass
class Invoice:
    \"\"\"청구 헤더 + 라인 모델. 청구 1건이 PO 여러 건을 묶을 수 있다(1:N).\"\"\"
    invoice_number: str
    vendor_code: str
    invoice_date: date
    po_numbers: list[str] = field(default_factory=list)
    items: list[InvoiceItem] = field(default_factory=list)

    @property
    def total_amount(self) -> Decimal:
        return sum(item.amount for item in self.items)
""")

commit("feat: define core data models")


# ─────────────────────────────────────────────
# Commit 3: feat - add config and data loader
# ─────────────────────────────────────────────
print("\n[3/12] feat: add config and data loader")

write("src/config.py", """\
\"\"\"프로젝트 설정값 모음.

검사 기준을 한 곳에 모아둔다. 이 값들만 바꾸면
코어 코드를 건드리지 않고 검사 강도를 조정할 수 있다.
\"\"\"
from decimal import Decimal

# ── 단가 정합성 ──────────────────────────────
# 청구 단가가 발주 단가 대비 몇 %까지 벗어나도 허용할지
PRICE_TOLERANCE_PERCENT = Decimal('3')

# ── 라운드 금액 탐지 ──────────────────────────
# 한도 회피가 의심되는 금액 목록. 청구 총액이 정확히 일치하면 적발.
ROUND_SUSPICIOUS_AMOUNTS = [
    Decimal('1000000'),    # 100만
    Decimal('5000000'),    # 500만
    Decimal('10000000'),   # 1000만
]

# ── 중복 청구 탐지 ────────────────────────────
# 같은 거래처·같은 금액 청구가 며칠 이내면 '근접 일자'로 보고 중복 의심
DUPLICATE_MAX_DAYS = 7

# ── 데이터 경로 ───────────────────────────────
VENDORS_CSV = "data/sample/vendors.csv"

# ── 거래처 상태 매핑 ──────────────────────────
# 공정위 CSV 표기 → Vendor 모델 표기
VENDOR_STATUS_MAP = {
    "정상영업": "정상",
    "폐업처리": "폐업",
    "휴업": "휴업",
}
""")

write("src/loader.py", """\
\"\"\"데이터 입력 계층: CSV를 읽어 도메인 객체로 변환한다.\"\"\"
import csv
import src.config as config
from src.models import Vendor


def load_vendors(csv_path):
    \"\"\"공정위 거래처 CSV를 읽어 Vendor 객체 리스트로 반환.

    CSV 컬럼: 상호 / 사업자등록번호 / 업소상태
    - 인코딩은 UTF-8-sig (한글 CSV의 BOM 처리)
    - vendor_code는 행 순서로 자동 생성 (V001, V002, ...)
    \"\"\"
    vendors = []
    with open(csv_path, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader, start=1):
            status = config.VENDOR_STATUS_MAP.get(row['업소상태'], "미확인")
            vendors.append(Vendor(
                vendor_code=f"V{i:03d}",
                business_number=row['사업자등록번호'],
                name=row['상호'],
                status=status,
            ))
    return vendors
""")

write("src/result.py", """\
from dataclasses import dataclass, field
from decimal import Decimal


@dataclass
class CheckResult:
    \"\"\"검사 한 건의 결과. Checker나 Detector가 검사할 때마다 하나씩 만든다.\"\"\"
    check_type: str         # "단가" / "수량" / "입고" / "중복" / "라운드" / "거래처"
    passed: bool            # True=정상, False=적발
    reason: str = ""        # 적발 사유 (통과면 빈 문자열)
    expected: Decimal = Decimal('0')
    actual: Decimal = Decimal('0')
    invoice_number: str = ""   # 탐지기가 채움


@dataclass
class MatchResult:
    \"\"\"청구(Invoice) 한 건에 대한 최종 판정.\"\"\"
    invoice_number: str
    status: str = "정상"
    checks: list[CheckResult] = field(default_factory=list)

    @property
    def reasons(self) -> list[str]:
        \"\"\"적발된 검사들의 사유만 모아서 반환.\"\"\"
        return [c.reason for c in self.checks if not c.passed]
""")

commit("feat: add config and data loader")


# ─────────────────────────────────────────────
# Commit 4: feat - implement tolerance rules (Strategy pattern)
# ─────────────────────────────────────────────
print("\n[4/12] feat: implement tolerance rules (Strategy pattern)")

write("src/rules/__init__.py", "")
write("src/rules/tolerance.py", """\
from abc import ABC, abstractmethod
from decimal import Decimal


class ToleranceRule(ABC):
    \"\"\"허용오차 판정 인터페이스 (Strategy 패턴).

    검증기는 이 룰에게 is_within(기준값, 실제값)만 물어본다.
    단가든 금액이든 동일한 인터페이스로 교체 가능.
    \"\"\"

    @abstractmethod
    def is_within(self, expected: Decimal, actual: Decimal) -> bool:
        ...


class PercentTolerance(ToleranceRule):
    \"\"\"퍼센트 기준 허용. 예: 3% → PercentTolerance(Decimal('3'))\"\"\"

    def __init__(self, percent: Decimal):
        self.percent = percent

    def is_within(self, expected: Decimal, actual: Decimal) -> bool:
        if expected == 0:
            return actual == 0
        diff = abs(actual - expected)
        limit = abs(expected) * self.percent / Decimal('100')
        return diff <= limit


class AbsoluteTolerance(ToleranceRule):
    \"\"\"절대값 기준 허용. 예: 100원 → AbsoluteTolerance(Decimal('100'))\"\"\"

    def __init__(self, amount: Decimal):
        self.amount = amount

    def is_within(self, expected: Decimal, actual: Decimal) -> bool:
        return abs(actual - expected) <= self.amount
""")

commit("feat: implement tolerance rules (Strategy pattern)")


# ─────────────────────────────────────────────
# Commit 5: feat - implement 3-way match consistency checker
# ─────────────────────────────────────────────
print("\n[5/12] feat: implement 3-way match consistency checker")

write("src/checkers/__init__.py", "")
write("src/checkers/consistency.py", """\
from src.result import CheckResult


class ConsistencyChecker:
    \"\"\"3-Way Match 정합성 검사 (단가·수량).

    - 단가: PO(발주 약속) vs Invoice(청구) — PercentTolerance 또는 AbsoluteTolerance 적용
    - 수량: GR(입고) vs Invoice(청구) — 허용오차 없음, 초과 시 즉시 적발
    \"\"\"

    def __init__(self, price_rule):
        self.price_rule = price_rule

    def check_price_values(self, po_price, invoice_price) -> CheckResult:
        \"\"\"발주단가(기준) vs 청구단가(실제) 비교.\"\"\"
        passed = self.price_rule.is_within(po_price, invoice_price)
        return CheckResult(
            check_type="단가",
            passed=passed,
            reason="" if passed else "단가차이",
            expected=po_price,
            actual=invoice_price,
        )

    def check_quantity_values(self, gr_qty, invoice_qty) -> CheckResult:
        \"\"\"입고량(기준) vs 청구량(실제) 비교. 청구량이 입고량을 초과하면 적발.\"\"\"
        passed = invoice_qty <= gr_qty
        return CheckResult(
            check_type="수량",
            passed=passed,
            reason="" if passed else "수량초과",
            expected=gr_qty,
            actual=invoice_qty,
        )
""")

commit("feat: implement 3-way match consistency checker")


# ─────────────────────────────────────────────
# Commit 6: feat - add anomaly detectors (duplicate, round-amount)
# ─────────────────────────────────────────────
print("\n[6/12] feat: add anomaly detectors (duplicate, round-amount)")

write("src/detectors/__init__.py", "")
write("src/detectors/base.py", """\
from abc import ABC, abstractmethod


class AnomalyDetector(ABC):
    \"\"\"부정 패턴 탐지기의 공통 뼈대 (추상 클래스).

    모든 탐지기는 detect()를 구현해 동일한 인터페이스를 따른다.
    (ConsistencyChecker의 check_*와 대칭되는 구조)
    \"\"\"

    @abstractmethod
    def detect(self, invoices) -> list:
        \"\"\"청구 리스트 전체를 받아, 적발된 건을 CheckResult 리스트로 반환.\"\"\"
        ...
""")

write("src/detectors/duplicate.py", """\
from decimal import Decimal
from src.detectors.base import AnomalyDetector
from src.result import CheckResult


class DuplicateDetector(AnomalyDetector):
    \"\"\"중복 청구 탐지.

    아래 세 조건이 모두 맞으면 중복으로 간주:
      1. 같은 거래처 (vendor_code 일치)
      2. 같은 총액 (total_amount 일치)
      3. 근접한 일자 (두 청구 날짜 차이가 max_days 이내)

    중복은 두 청구 모두 보류가 되어야 하므로 a·b 각각 CheckResult를 생성한다.
    \"\"\"

    def __init__(self, max_days: int):
        self.max_days = max_days

    def detect(self, invoices) -> list:
        results = []
        for i in range(len(invoices)):
            for j in range(i + 1, len(invoices)):
                a, b = invoices[i], invoices[j]
                same_vendor = a.vendor_code == b.vendor_code
                same_amount = a.total_amount == b.total_amount
                day_gap = abs((a.invoice_date - b.invoice_date).days)
                if same_vendor and same_amount and day_gap <= self.max_days:
                    reason = f"중복 청구 의심: {a.invoice_number}와 {b.invoice_number}"
                    for num in (a.invoice_number, b.invoice_number):
                        results.append(CheckResult(
                            check_type="중복",
                            passed=False,
                            reason=reason,
                            expected=Decimal('0'),
                            actual=b.total_amount,
                            invoice_number=num,
                        ))
        return results
""")

write("src/detectors/round_amount.py", """\
from decimal import Decimal
from src.detectors.base import AnomalyDetector
from src.result import CheckResult


class RoundAmountDetector(AnomalyDetector):
    \"\"\"라운드 금액 탐지.

    정상 거래는 단가×수량이라 금액이 지저분하지만(예: 16,211원),
    결재 한도를 피하려는 조작이면 깔끔한 숫자가 나온다(예: 정확히 5,000,000원).
    의심 금액 목록은 config.py에서 주입받는다.
    \"\"\"

    def __init__(self, suspicious_amounts: list):
        self.suspicious_amounts = suspicious_amounts

    def detect(self, invoices) -> list:
        results = []
        for invoice in invoices:
            total = invoice.total_amount
            if total in self.suspicious_amounts:
                results.append(CheckResult(
                    check_type="라운드",
                    passed=False,
                    reason="라운드 금액(한도 회피 의심)",
                    expected=Decimal('0'),
                    actual=total,
                    invoice_number=invoice.invoice_number,
                ))
        return results
""")

commit("feat: add anomaly detectors (duplicate, round-amount)")


# ─────────────────────────────────────────────
# Commit 7: feat - add vendor verifier with NTS API integration
# ─────────────────────────────────────────────
print("\n[7/12] feat: add vendor verifier with NTS API integration")

write("src/verifiers/__init__.py", "")
write("src/verifiers/base.py", """\
\"\"\"거래처 실재성 검증의 추상 뼈대.

모든 거래처 검증기는 이 인터페이스를 상속해 verify()를 구현한다.
AnomalyDetector(추상)를 detect()로 통일한 것과 같은 Strategy 패턴.
\"\"\"
from abc import ABC, abstractmethod


class VendorVerifier(ABC):
    \"\"\"거래처 검증기 인터페이스.

    반환값은 Vendor.VALID_STATUSES 중 하나:
        "정상" / "폐업" / "휴업" / "검증불가"
    \"\"\"

    @abstractmethod
    def verify(self, business_number: str) -> str:
        \"\"\"사업자번호(하이픈 없는 10자리)를 받아 상태 문자열을 반환.\"\"\"
        ...
""")

write("src/verifiers/api.py", """\
\"\"\"국세청 사업자등록 상태조회 API 기반 거래처 검증기.

API 키는 생성자로 주입받는다 (.env 분리 원칙).
API 호출 실패 시 '검증불가'로 폴백하여 코어가 죽지 않게 한다.
\"\"\"
import requests
from src.verifiers.base import VendorVerifier


class APIVendorVerifier(VendorVerifier):
    \"\"\"국세청 공공 API로 사업자 상태를 조회하는 검증기.\"\"\"

    URL = "https://api.odcloud.kr/api/nts-businessman/v1/status"

    # 국세청 API 상태 문자열 → 내부 표기 매핑
    STATUS_MAP = {
        "계속사업자": "정상",
        "휴업자": "휴업",
        "폐업자": "폐업",
    }

    def __init__(self, service_key: str, timeout: int = 5):
        self.service_key = service_key
        self.timeout = timeout

    def verify(self, business_number: str) -> str:
        \"\"\"사업자번호 1건을 조회해 상태 문자열을 반환.\"\"\"
        try:
            res = requests.post(
                self.URL,
                params={"serviceKey": self.service_key},
                json={"b_no": [business_number]},
                headers={"Content-Type": "application/json"},
                timeout=self.timeout,
            )
            if res.status_code != 200:
                return "검증불가"
            items = res.json().get("data", [])
            if not items:
                return "검증불가"
            return self.STATUS_MAP.get(items[0]["b_stt"], "검증불가")
        except Exception:
            return "검증불가"
""")

commit("feat: add vendor verifier with NTS API integration")


# ─────────────────────────────────────────────
# Commit 8: feat - wire up detection engine
# ─────────────────────────────────────────────
print("\n[8/12] feat: wire up detection engine")

write("src/engine.py", """\
\"\"\"3축 검증을 지휘하는 핵심 엔진.

3축 = (1) 거래 정합성  (2) 부정 패턴 탐지  (3) 거래처 실재성 검증
\"\"\"
from decimal import Decimal
from src.result import CheckResult, MatchResult


class MatchingEngine:

    def __init__(self, consistency_checker, detectors):
        \"\"\"
        consistency_checker: ConsistencyChecker
        detectors: AnomalyDetector 리스트
        \"\"\"
        self.checker = consistency_checker
        self.detectors = detectors

    def _empty_row(self) -> dict:
        return {
            "po_qty": Decimal('0'), "po_price": None,
            "gr_qty": Decimal('0'),
            "inv_qty": Decimal('0'), "inv_price": None,
        }

    def aggregate(self, invoice, purchase_orders, goods_receipts) -> dict:
        \"\"\"청구 1건 기준으로 자재별 PO·GR·Invoice 수량/단가를 합산한다.\"\"\"
        agg = {}

        # 1) 이 청구가 가리키는 PO만 추출 (1:N 지원)
        related_pos = [po for po in purchase_orders if po.po_number in invoice.po_numbers]

        # 2) PO 라인 합산
        for po in related_pos:
            for item in po.items:
                if item.material_code not in agg:
                    agg[item.material_code] = self._empty_row()
                agg[item.material_code]["po_qty"] += item.quantity
                agg[item.material_code]["po_price"] = item.unit_price

        # 3) GR 라인 합산
        related_po_numbers = {po.po_number for po in related_pos}
        for gr in goods_receipts:
            if gr.po_number in related_po_numbers:
                for item in gr.items:
                    if item.material_code not in agg:
                        agg[item.material_code] = self._empty_row()
                    agg[item.material_code]["gr_qty"] += item.quantity

        # 4) Invoice 라인 합산
        for item in invoice.items:
            if item.material_code not in agg:
                agg[item.material_code] = self._empty_row()
            agg[item.material_code]["inv_qty"] += item.quantity
            agg[item.material_code]["inv_price"] = item.unit_price

        return agg

    def check_invoice(self, invoice, purchase_orders, goods_receipts) -> list:
        \"\"\"청구 1건의 정합성(단가·수량·입고유무)을 검사해 CheckResult 리스트 반환.\"\"\"
        agg = self.aggregate(invoice, purchase_orders, goods_receipts)
        checks = []

        for material, row in agg.items():
            # (1) 입고기록 없음
            if row["inv_qty"] > 0 and row["gr_qty"] == 0:
                checks.append(CheckResult(
                    check_type="입고",
                    passed=False,
                    reason="입고기록 없음",
                    expected=Decimal('0'),
                    actual=row["inv_qty"],
                ))
                continue

            # (2) 단가 검사
            if row["po_price"] is not None and row["inv_price"] is not None:
                checks.append(self.checker.check_price_values(row["po_price"], row["inv_price"]))

            # (3) 수량 검사
            checks.append(self.checker.check_quantity_values(row["gr_qty"], row["inv_qty"]))

        return checks

    def run(self, invoices, purchase_orders, goods_receipts) -> dict:
        \"\"\"전체 청구를 판정해 invoice_number → MatchResult 딕셔너리로 반환.\"\"\"
        results = {}

        # 1) 정합성 검사
        for invoice in invoices:
            checks = self.check_invoice(invoice, purchase_orders, goods_receipts)
            results[invoice.invoice_number] = MatchResult(
                invoice_number=invoice.invoice_number,
                checks=checks,
            )

        # 2) 부정 패턴 탐지
        for detector in self.detectors:
            for cr in detector.detect(invoices):
                if cr.invoice_number in results:
                    results[cr.invoice_number].checks.append(cr)

        # 3) 최종 status 확정: 하나라도 적발이면 보류
        for mr in results.values():
            if any(not c.passed for c in mr.checks):
                mr.status = "보류"

        return results
""")

commit("feat: wire up detection engine")


# ─────────────────────────────────────────────
# Commit 9: feat - add result aggregation and Excel report
# ─────────────────────────────────────────────
print("\n[9/12] feat: add result aggregation and Excel report")

write("src/report.py", """\
\"\"\"판정 결과를 엑셀 리포트로 출력한다 (openpyxl).

- audit_report.xlsx : 판정결과 시트 + 요약 시트
- source_data.xlsx  : 원본 거래 데이터 (발주/입고/청구 시트)
\"\"\"
import openpyxl
from openpyxl.styles import PatternFill, Font


class ReportGenerator:
    \"\"\"MatchResult 모음을 받아 엑셀 감사 리포트를 만든다.\"\"\"

    HOLD_FILL = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    HEADER_FONT = Font(bold=True)

    def __init__(self, results, invoices, vendor_by_code):
        self.results = results
        self.invoices = invoices
        self.vendor_by_code = vendor_by_code

    def generate(self, path="output/audit_report.xlsx") -> str:
        \"\"\"판정 결과 엑셀을 만들어 저장한다.\"\"\"
        import os
        os.makedirs(os.path.dirname(path), exist_ok=True)
        wb = openpyxl.Workbook()
        self._add_result_sheet(wb)
        self._add_summary_sheet(wb)
        wb.save(path)
        return path

    def _add_result_sheet(self, wb):
        ws = wb.active
        ws.title = "판정결과"
        ws.append(["청구번호", "거래처", "상태", "사유"])
        for cell in ws[1]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT

        for invoice in self.invoices:
            mr = self.results[invoice.invoice_number]
            vendor = self.vendor_by_code[invoice.vendor_code]
            ws.append([
                mr.invoice_number,
                vendor.name,
                mr.status,
                ", ".join(mr.reasons) if mr.reasons else "-",
            ])
            if mr.status == "보류":
                for cell in ws[ws.max_row]:
                    cell.fill = self.HOLD_FILL

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 45

    def _add_summary_sheet(self, wb):
        ws = wb.create_sheet("요약")
        total = len(self.invoices)
        normal = sum(1 for inv in self.invoices if self.results[inv.invoice_number].status == "정상")
        hold = total - normal

        ws.append(["구분", "건수"])
        ws.append(["총 검증", total])
        ws.append(["정상", normal])
        ws.append(["보류", hold])
        ws.append([])

        sub_header_row = ws.max_row + 1
        ws.append(["사유별 분포", "건수"])
        reason_count = {}
        for inv in self.invoices:
            for check in self.results[inv.invoice_number].checks:
                if not check.passed:
                    reason_count[check.check_type] = reason_count.get(check.check_type, 0) + 1
        for key, cnt in reason_count.items():
            ws.append([key, cnt])

        for cell in ws[1]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
        for cell in ws[sub_header_row]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 8

    def generate_source(self, pos, grs, invs, path="output/source_data.xlsx") -> str:
        \"\"\"원본 거래 데이터(발주/입고/청구)를 3개 시트로 출력한다.\"\"\"
        import os
        os.makedirs(os.path.dirname(path), exist_ok=True)
        wb = openpyxl.Workbook()

        ws_po = wb.active
        ws_po.title = "발주(PO)"
        ws_po.append(["발주번호", "거래처코드", "발주일", "품목", "수량", "단가", "금액"])
        for po in pos:
            for item in po.items:
                ws_po.append([po.po_number, po.vendor_code, po.order_date,
                               item.material_code, item.quantity, item.unit_price, item.amount])

        ws_gr = wb.create_sheet("입고(GR)")
        ws_gr.append(["입고번호", "발주번호", "입고일", "품목", "수량"])
        for gr in grs:
            for item in gr.items:
                ws_gr.append([gr.gr_number, gr.po_number, gr.receipt_date,
                               item.material_code, item.quantity])

        ws_inv = wb.create_sheet("청구(Invoice)")
        ws_inv.append(["청구번호", "거래처코드", "청구일", "발주번호들", "품목", "수량", "단가", "금액"])
        for inv in invs:
            po_text = ", ".join(inv.po_numbers)
            for item in inv.items:
                ws_inv.append([inv.invoice_number, inv.vendor_code, inv.invoice_date,
                                po_text, item.material_code, item.quantity, item.unit_price, item.amount])

        for ws in [ws_po, ws_gr, ws_inv]:
            for cell in ws[1]:
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT

        wb.save(path)
        return path
""")

commit("feat: add result aggregation and Excel report")


# ─────────────────────────────────────────────
# Commit 10: feat - add entrypoint and sample data generator
# ─────────────────────────────────────────────
print("\n[10/12] feat: add entrypoint and sample data generator")

write("scripts/__init__.py", "")
write("scripts/sample_data.py", """\
\"\"\"합성 거래 데이터 생성 (9케이스).

거래처는 실데이터(vendors.csv)지만, 주문·입고·청구는
기업 내부 데이터라 공개되지 않으므로 실무 시나리오대로 합성한다.
\"\"\"
from decimal import Decimal
from datetime import date
from src.models import PurchaseOrder, POItem, GoodsReceipt, GRItem, Invoice, InvoiceItem


def build_sample_transactions(vendors):
    \"\"\"거래처 리스트를 받아 거래(PO/GR/Invoice)를 생성해 반환.

    반환: (purchase_orders, goods_receipts, invoices)
    \"\"\"
    normal_vendors = [v for v in vendors if v.status == "정상"]
    closed_vendors = [v for v in vendors if v.status == "폐업"]

    pos, grs, invs = [], [], []

    def add(no, vendor, po_items, gr_items, inv_items, inv_date=date(2025, 6, 10)):
        pos.append(PurchaseOrder(
            f"PO-{no}", vendor.vendor_code, date(2025, 6, 1),
            items=[POItem(m, Decimal(str(q)), Decimal(str(p))) for m, q, p in po_items]
        ))
        if gr_items is not None:
            grs.append(GoodsReceipt(
                f"GR-{no}", f"PO-{no}", date(2025, 6, 3),
                items=[GRItem(m, Decimal(str(q))) for m, q in gr_items]
            ))
        invs.append(Invoice(
            f"INV-{no}", vendor.vendor_code, inv_date,
            po_numbers=[f"PO-{no}"],
            items=[InvoiceItem(m, Decimal(str(q)), Decimal(str(p))) for m, q, p in inv_items]
        ))

    nv = lambda i: normal_vendors[i % len(normal_vendors)]
    cv = lambda i: closed_vendors[i % len(closed_vendors)]

    # 케이스1: 다 일치 → 정상
    add(1, nv(0), [("M1", 10, 1000)], [("M1", 10)], [("M1", 10, 1000)])
    # 케이스2: 단가 5%↑ (허용 3%) → 단가차이
    add(2, nv(1), [("M1", 10, 1100)], [("M1", 10)], [("M1", 10, 1155)])
    # 케이스3: 청구수량 > 입고수량 → 수량초과
    add(3, nv(2), [("M1", 10, 1200)], [("M1", 10)], [("M1", 12, 1200)])
    # 케이스4: 부분입고 후 전량청구 → 수량초과
    add(4, nv(3), [("M1", 10, 1300)], [("M1", 7)], [("M1", 10, 1300)])
    # 케이스5: 입고기록 없음
    add(5, nv(4), [("M1", 10, 1400)], None, [("M1", 10, 1400)])
    # 케이스6: 단가 1% (허용 3%) → 정상
    add(6, nv(5), [("M1", 10, 1500)], [("M1", 10)], [("M1", 10, 1515)])
    # 케이스7: 정확히 500만 → 라운드 금액
    add(7, nv(6), [("M1", 1, 5000000)], [("M1", 1)], [("M1", 1, 5000000)])
    # 케이스8: 같은 거래처·금액·근접일자 2건 → 중복 청구
    dup_vendor = nv(7)
    add("8a", dup_vendor, [("M1", 1, 3000)], [("M1", 1)], [("M1", 1, 3000)], date(2025, 6, 10))
    add("8b", dup_vendor, [("M1", 1, 3000)], [("M1", 1)], [("M1", 1, 3000)], date(2025, 6, 12))
    # 케이스9: 정합성 정상이지만 거래처 폐업 → 거래처 검증 축에서 적발
    add(9, cv(0), [("M1", 10, 2000)], [("M1", 10)], [("M1", 10, 2000)])

    return pos, grs, invs
""")

write("scripts/main.py", """\
\"\"\"매입거래 부정·이상 탐지 엔진 — 진입점.

흐름: 설정 읽기 → 거래처 로딩 → 거래 합성 → 엔진 조립 → 실행 → 결과 출력
\"\"\"
import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dotenv import load_dotenv
from src import config
from src.loader import load_vendors
from src.rules.tolerance import PercentTolerance
from src.checkers.consistency import ConsistencyChecker
from src.detectors.duplicate import DuplicateDetector
from src.detectors.round_amount import RoundAmountDetector
from src.engine import MatchingEngine
from src.result import CheckResult
from src.report import ReportGenerator
from src.verifiers.api import APIVendorVerifier
from scripts.sample_data import build_sample_transactions


def main():
    # 1) 거래처 로딩 (공정위 실데이터)
    vendors = load_vendors(config.VENDORS_CSV)
    print(f"거래처 {len(vendors)}개 로딩 완료 "
          f"(정상 {sum(1 for v in vendors if v.status=='정상')}, "
          f"폐업 {sum(1 for v in vendors if v.status=='폐업')})")

    # 2) 거래 데이터 합성
    pos, grs, invs = build_sample_transactions(vendors)
    print(f"거래 합성: 발주 {len(pos)} / 입고 {len(grs)} / 청구 {len(invs)}\\n")

    # 3) 엔진 조립
    checker = ConsistencyChecker(PercentTolerance(config.PRICE_TOLERANCE_PERCENT))
    detectors = [
        RoundAmountDetector(config.ROUND_SUSPICIOUS_AMOUNTS),
        DuplicateDetector(config.DUPLICATE_MAX_DAYS),
    ]
    engine = MatchingEngine(checker, detectors)

    # 4) 거래처 검증기 조립
    load_dotenv()
    api_verifier = APIVendorVerifier(os.getenv("NTS_SERVICE_KEY"))
    vendor_by_code = {v.vendor_code: v for v in vendors}

    # 5) 엔진 실행
    results = engine.run(invs, pos, grs)

    # 6) 거래처 실재성 검증 (국세청 API) — 엔진 판정 후 외부 재검증
    for invoice in invs:
        vendor = vendor_by_code[invoice.vendor_code]
        try:
            status = api_verifier.verify(vendor.business_number)
        except Exception:
            status = "검증불가"

        passed = status == "정상"
        vendor_check = CheckResult(
            check_type="거래처",
            passed=passed,
            reason="" if passed else f"거래처 상태: {status}",
        )
        results[invoice.invoice_number].checks.append(vendor_check)
        if not passed:
            results[invoice.invoice_number].status = "보류"

    # 7) 결과 출력
    print("=" * 55)
    print("판정 결과")
    print("=" * 55)
    normal_count = sum(1 for inv in invs if results[inv.invoice_number].status == "정상")
    hold_count = len(invs) - normal_count

    for invoice in invs:
        mr = results[invoice.invoice_number]
        reasons = ", ".join(mr.reasons) if mr.reasons else "-"
        print(f"  {mr.invoice_number:8s} [{mr.status}]  {reasons}")

    print("\\n" + "-" * 55)
    print(f"총 {len(invs)}건  |  정상 {normal_count}  |  보류 {hold_count}")

    # 8) 엑셀 리포트 출력
    report = ReportGenerator(results, invs, vendor_by_code)
    print(f"\\n엑셀 리포트 생성: {report.generate()}")
    print(f"원본 데이터 생성: {report.generate_source(pos, grs, invs)}")


if __name__ == "__main__":
    main()
""")

commit("feat: add entrypoint and sample data generator")


# ─────────────────────────────────────────────
# Commit 11: docs - add sample vendor data
# ─────────────────────────────────────────────
print("\n[11/12] docs: add sample vendor data")

write("data/sample/vendors.csv", """\
상호,사업자등록번호,업소상태
주식회사가나다,1234567890,정상영업
테스트상사,9876543210,정상영업
한국물산,1111111111,정상영업
서울유통,2222222222,정상영업
부산물류,3333333333,정상영업
대구상사,4444444444,정상영업
인천무역,5555555555,정상영업
광주기업,6666666666,정상영업
대전산업,7777777777,폐업처리
울산제조,8888888888,폐업처리
""")

commit("docs: add sample vendor data")


# ─────────────────────────────────────────────
# Commit 12: docs - update README
# ─────────────────────────────────────────────
print("\n[12/12] docs: update README with architecture and usage")

write("README.md", """\
# purchase-fraud-detector

ERP가 통과시킨 매입거래를 **3축 병렬 검증**으로 재검증해 부정·이상을 잡아내는 외부 감사 분석 도구입니다.

---

## 프로젝트 소개

SAP·GRC 모듈을 갖춘 대기업과 달리, 영림원·더존·이카운트를 사용하는 중소·중견 기업은 **부정 패턴 탐지나 외부 공공 API 거래처 검증** 같은 상위 감사 레이어를 갖추기 어렵습니다.

이 프로젝트는 **SAP Clean Core의 Side-by-Side 외부 확장 패턴**에 따라, ERP에서 추출한 거래 데이터에 다중 신호 분석 레이어를 얹는 구조입니다. ERP 종류와 무관하게 이미 통과된 거래를 **독립적으로 재검증**합니다.

---

## 주요 기능 (3축 병렬 검증)

### ① 거래 정합성 검증 (3-Way Match)
PO(발주) · GR(입고) · Invoice(청구)의 수량·단가를 자동 대조합니다.

- **단가 검사:** 발주 단가 대비 청구 단가가 허용오차(기본 3%)를 벗어나면 적발
- **수량 검사:** 청구 수량이 입고 수량을 초과하면 적발
- **입고기록 검사:** 입고 없이 청구만 도착한 건 적발
- 청구 1건이 PO 여러 건을 묶는 **1:N 매칭**, 부분입고·분할청구 지원

허용오차는 Strategy 패턴(`PercentTolerance` / `AbsoluteTolerance`)으로 교체 가능합니다.

### ② 부정 패턴 탐지
- **중복 청구 탐지:** 같은 거래처가 7일 이내에 같은 금액을 두 번 청구하면 적발
- **라운드 금액 탐지:** 결재 한도 회피가 의심되는 금액(정확히 500만 원 등)이면 적발

### ③ 거래처 실재성 검증
**국세청 사업자등록 상태조회 API**를 실시간 호출해 폐업·휴업 거래처를 적발합니다.

---

## 판정 흐름

```
1. 입력      거래처 마스터(공정위 실데이터) + 거래 3종(합성)
       │
2. 전처리    데이터 로딩 · PO 기준 집계 (1:N · 부분입고 · 분할청구)
       │
3. 검증·탐지 [정합성] [부정 패턴] [거래처 실재성: 국세청 API]  ← 3축
       │
4. 판정      통합 판정 · 사유 자동 분류 (하나라도 적발 → 보류)
       │
5. 출력      엑셀 리포트 (판정결과 · 요약 통계 + 원본 데이터)
```

---

## 프로젝트 구조

```
purchase-fraud-detector/
├── src/                        # 핵심 패키지
│   ├── models.py               # 도메인 모델 (PO·GR·Invoice·Vendor)
│   ├── config.py               # 검사 기준 설정값
│   ├── loader.py               # CSV → 도메인 객체 변환
│   ├── result.py               # CheckResult · MatchResult
│   ├── engine.py               # 3축 검증 엔진
│   ├── report.py               # 엑셀 리포트 생성
│   ├── rules/
│   │   └── tolerance.py        # 허용오차 Strategy (Percent / Absolute)
│   ├── checkers/
│   │   └── consistency.py      # 3-Way Match 정합성 검사
│   ├── detectors/
│   │   ├── base.py             # AnomalyDetector 추상 클래스
│   │   ├── duplicate.py        # 중복 청구 탐지
│   │   └── round_amount.py     # 라운드 금액 탐지
│   └── verifiers/
│       ├── base.py             # VendorVerifier 추상 클래스
│       └── api.py              # 국세청 API 검증기
├── scripts/
│   ├── main.py                 # 실행 진입점
│   └── sample_data.py          # 합성 거래 데이터 생성 (9케이스)
├── data/
│   └── sample/
│       └── vendors.csv         # 샘플 거래처 데이터
├── output/                     # 생성 리포트 (gitignore)
├── tests/                      # 테스트 (추후 추가)
├── .env.example                # API 키 환경변수 템플릿
├── requirements.txt
└── README.md
```

---

## 빠른 시작

```bash
# 1. 클론
git clone https://github.com/dabok0519/purchase-fraud-detector.git
cd purchase-fraud-detector

# 2. 의존성 설치
pip install -r requirements.txt

# 3. API 키 설정 (선택)
cp .env.example .env
# .env 파일에 NTS_SERVICE_KEY=발급받은키 입력

# 4. 실행
python scripts/main.py
```

> API 키가 없어도 정합성·부정 패턴 검증은 정상 동작합니다. 거래처 검증만 '검증불가'로 처리됩니다.

---

## 설계 원칙

- **결정론적 코어 우선:** 외부 의존성 없이 합성 데이터만으로 모든 판정이 동작
- **확장 가능한 검증 구조:** 검사·탐지 클래스가 서로 독립적이라 새 신호 추가가 쉬움
- **외부 의존성 격리:** 거래처 검증을 추상 인터페이스 + 폴백으로 설계해 견고성 확보

---

## 기술 스택

- **Python** — OOP 설계 (dataclass, ABC/abstractmethod, Strategy 패턴, 의존성 주입)
- **requests** — 국세청 API 호출
- **openpyxl** — 엑셀 리포트 출력
- **python-dotenv** — API 키 환경변수 분리
""")

commit("docs: update README with architecture and usage")

print("\n✅ 완료! 12개 commit이 생성되었습니다.")
print("   git log --oneline 으로 확인하세요.")
print("\n   push:")
print("   git push origin main")